import logging
import os
from pathlib import Path

from openpyxl import load_workbook
from dotenv import load_dotenv


load_dotenv()
logger = logging.getLogger(__name__)


def save_formated_result_1c(result_list, file_name):
    """
    Save formatted file from template wb
    :param result_list:
    :param file_name:
    :return:
    """
    file_template_path = os.getenv('path_template_1c')
    wb = load_workbook(file_template_path)
    ws = wb.worksheets[0]

    # B-столбец (1 = A, 2 = B, 3 = C, ...)
    column_vendor_code_index = int(os.getenv('column_vendor_code_index'))
    column_description_index = int(os.getenv('column_description_index'))
    start_row = int(os.getenv('start_row'))

    for i, result in enumerate(result_list, start=start_row):
        vendor_code = result.get('vendor_code')
        text_description = result.get('text_description')

        ws.cell(row=i, column=column_vendor_code_index, value=vendor_code)
        ws.cell(row=i, column=column_description_index, value=text_description)

    full_path_result = Path(os.getenv('sub_path_format_result_wb')) / file_name
    wb.save(full_path_result)
    wb.close()
    logger.info(f'Файл - {file_name} успешно сохранен')

def save_formated_result_wb(result_list, file_name):
    """
    Save formatted file from template wb
    :param result_list:
    :param file_name:
    :return:
    """
    file_template_path = os.getenv('path_template')
    wb = load_workbook(file_template_path)
    ws = wb[os.getenv('active_sheet')]

    # B-столбец (1 = A, 2 = B, 3 = C, ...)
    column_vendor_code_index = int(os.getenv('column_vendor_code_index'))
    column_description_index = int(os.getenv('column_description_index'))
    start_row = int(os.getenv('start_row'))

    for i, result in enumerate(result_list, start=start_row):
        vendor_code = result.get('vendor_code')
        text_description = result.get('text_description')

        ws.cell(row=i, column=column_vendor_code_index, value=vendor_code)
        ws.cell(row=i, column=column_description_index, value=text_description)

    full_path_result = Path(os.getenv('sub_path_format_result_wb')) / file_name
    wb.save(full_path_result)
    wb.close()
    logger.info(f'Файл - {file_name} успешно сохранен')

